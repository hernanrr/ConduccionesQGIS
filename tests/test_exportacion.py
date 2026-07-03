"""Pruebas unitarias para el nucleo reusable de exportacion."""

from __future__ import annotations

import csv

import pytest
from ConduccionesQGIS.core.alineamiento import construir_tabla_alineamiento
from ConduccionesQGIS.core.exportacion import (
    COLUMNAS_PERFIL_PUNTOS,
    COLUMNAS_PERFIL_TRAMOS,
    COLUMNAS_PUNTOS_INTERSECCION,
    COLUMNAS_TRAMOS,
    FORMATO_CSV,
    FORMATO_XLSX,
    NOMBRE_TABLA_PERFIL_PUNTOS,
    NOMBRE_TABLA_PERFIL_TRAMOS,
    NOMBRE_TABLA_PUNTOS_INTERSECCION,
    NOMBRE_TABLA_TRAMOS,
    construir_ruta_exportacion,
    crear_tabla_exportable,
    exportar_tabla_csv,
    exportar_tablas,
    exportar_tablas_csv,
    exportar_tablas_xlsx,
    serializar_tablas_alineamiento,
    serializar_tablas_perfil,
    validar_nombre_base,
)
from ConduccionesQGIS.core.models import Punto2D
from ConduccionesQGIS.core.perfil import construir_perfil_longitudinal
from openpyxl import load_workbook


def test_crear_tabla_exportable_preserva_columnas_y_filas() -> None:
    """Construye una tabla exportable estable a partir de secuencias simples."""
    tabla = crear_tabla_exportable(
        "puntos_interseccion",
        ["PI", "Progresiva"],
        [(1, 0.0), (2, 125.5)],
    )

    assert tabla.nombre == "puntos_interseccion"
    assert tabla.columnas == ("PI", "Progresiva")
    assert tabla.filas == ((1, 0.0), (2, 125.5))


def test_validar_nombre_base_rechaza_vacio() -> None:
    """Exige que el nombre base manual tenga contenido."""
    with pytest.raises(ValueError, match="obligatorio"):
        validar_nombre_base("   ")


def test_validar_nombre_base_rechaza_caracteres_invalidos() -> None:
    """Bloquea caracteres no validos para nombres de archivo en Windows."""
    with pytest.raises(ValueError, match="no permitidos"):
        validar_nombre_base("alineamiento:demo")


def test_construir_ruta_exportacion_anexa_nombre_y_extension(tmp_path) -> None:
    """Deriva la ruta final desde directorio, base, tabla y extension."""
    ruta = construir_ruta_exportacion(
        tmp_path,
        "alineamiento_demo",
        "puntos_interseccion",
        "csv",
    )

    assert ruta.name == "alineamiento_demo_puntos_interseccion.csv"
    assert ruta.parent == tmp_path


def test_exportar_tabla_csv_genera_encabezados_y_filas(tmp_path) -> None:
    """Exporta una tabla individual a CSV con encabezado y contenido."""
    tabla = crear_tabla_exportable(
        "puntos_interseccion",
        ["PI", "Progresiva", "Giro"],
        [(1, 0.0, "Inicio"), (2, 125.5, "Derecha")],
    )

    ruta = exportar_tabla_csv(tabla, tmp_path, "alineamiento_demo")

    assert ruta.exists()
    with ruta.open("r", encoding="utf-8-sig", newline="") as descriptor:
        filas = list(csv.reader(descriptor))

    assert filas == [
        ["PI", "Progresiva", "Giro"],
        ["1", "0.0", "Inicio"],
        ["2", "125.5", "Derecha"],
    ]


def test_exportar_tablas_csv_genera_un_archivo_por_tabla(tmp_path) -> None:
    """Exporta varias tablas a CSV separados por nombre logico."""
    puntos = crear_tabla_exportable("puntos_interseccion", ["PI"], [(1,), (2,)])
    tramos = crear_tabla_exportable("tramos", ["Tramo"], [(1,), (2,)])

    rutas = exportar_tablas_csv([puntos, tramos], tmp_path, "alineamiento_demo")

    assert set(rutas) == {"puntos_interseccion", "tramos"}
    assert rutas["puntos_interseccion"].name.endswith("_puntos_interseccion.csv")
    assert rutas["tramos"].name.endswith("_tramos.csv")


def test_serializar_tablas_alineamiento_expone_contrato_estable() -> None:
    """Serializa puntos y tramos con nombres y columnas fijos para v0.2."""
    vertices = [
        Punto2D(0.0, 0.0),
        Punto2D(0.0, 100.0),
        Punto2D(100.0, 100.0),
        Punto2D(100.0, 200.0),
    ]
    puntos, tramos = construir_tabla_alineamiento(vertices)

    tabla_puntos, tabla_tramos = serializar_tablas_alineamiento(puntos, tramos)

    assert tabla_puntos.nombre == NOMBRE_TABLA_PUNTOS_INTERSECCION
    assert tabla_puntos.columnas == COLUMNAS_PUNTOS_INTERSECCION
    assert len(tabla_puntos.filas) == len(puntos)

    assert tabla_tramos.nombre == NOMBRE_TABLA_TRAMOS
    assert tabla_tramos.columnas == COLUMNAS_TRAMOS
    assert len(tabla_tramos.filas) == len(tramos)


def test_serializar_tablas_alineamiento_mantiene_coherencia_entre_pi_y_tramos() -> None:
    """Mantiene coherencia entre progresivas, tramos y referencias de PI."""
    vertices = [
        Punto2D(0.0, 0.0),
        Punto2D(0.0, 100.0),
        Punto2D(100.0, 100.0),
        Punto2D(100.0, 200.0),
    ]
    puntos, tramos = construir_tabla_alineamiento(vertices)

    tabla_puntos, tabla_tramos = serializar_tablas_alineamiento(puntos, tramos)

    primera_fila_tramo = tabla_tramos.filas[0]
    segunda_fila_punto = tabla_puntos.filas[1]

    assert primera_fila_tramo[1] == 1
    assert primera_fila_tramo[2] == 2
    assert primera_fila_tramo[3] == pytest.approx(100.0)
    assert segunda_fila_punto[1] == pytest.approx(100.0)


def test_exportar_tablas_xlsx_genera_hojas_separadas(tmp_path) -> None:
    """Exporta varias tablas a un libro XLSX con una hoja por tabla."""
    puntos = crear_tabla_exportable("puntos_interseccion", ["PI"], [(1,), (2,)])
    tramos = crear_tabla_exportable("tramos", ["Tramo"], [(1,), (2,)])

    ruta = exportar_tablas_xlsx([puntos, tramos], tmp_path, "alineamiento_demo")

    assert ruta.exists()
    libro = load_workbook(ruta)
    assert libro.sheetnames == ["puntos_interseccion", "tramos"]
    assert libro["puntos_interseccion"]["A1"].value == "PI"
    assert libro["puntos_interseccion"]["A2"].value == 1
    assert libro["tramos"]["A2"].value == 1


def test_exportar_tablas_permite_csv_y_xlsx_en_misma_ejecucion(tmp_path) -> None:
    """Genera todos los formatos solicitados en una unica llamada."""
    puntos = crear_tabla_exportable("puntos_interseccion", ["PI"], [(1,), (2,)])
    tramos = crear_tabla_exportable("tramos", ["Tramo"], [(1,), (2,)])

    resultado = exportar_tablas(
        [puntos, tramos],
        tmp_path,
        "alineamiento_demo",
        [FORMATO_CSV, FORMATO_XLSX],
    )

    assert set(resultado.csv) == {"puntos_interseccion", "tramos"}
    assert resultado.xlsx is not None and resultado.xlsx.exists()


def test_serializar_tablas_perfil_expone_contrato_estable() -> None:
    """Serializa el perfil longitudinal con nombres y columnas fijos."""
    vertices = [
        Punto2D(0.0, 0.0),
        Punto2D(0.0, 100.0),
        Punto2D(100.0, 100.0),
    ]
    puntos, tramos = construir_tabla_alineamiento(vertices)
    perfil_puntos, perfil_tramos = construir_perfil_longitudinal(
        puntos,
        tramos,
        [10.0, 12.5, 15.0],
    )

    tabla_puntos, tabla_tramos = serializar_tablas_perfil(
        perfil_puntos,
        perfil_tramos,
    )

    assert tabla_puntos.nombre == NOMBRE_TABLA_PERFIL_PUNTOS
    assert tabla_puntos.columnas == COLUMNAS_PERFIL_PUNTOS
    assert tabla_puntos.filas[1] == (2, pytest.approx(100.0), pytest.approx(12.5))

    assert tabla_tramos.nombre == NOMBRE_TABLA_PERFIL_TRAMOS
    assert tabla_tramos.columnas == COLUMNAS_PERFIL_TRAMOS
    assert len(tabla_tramos.filas) == len(perfil_tramos)


def test_serializar_tablas_perfil_mantiene_delta_z_y_pendiente() -> None:
    """Conserva los calculos del perfil al serializar las tablas exportables."""
    vertices = [
        Punto2D(0.0, 0.0),
        Punto2D(30.0, 40.0),
        Punto2D(60.0, 40.0),
    ]
    puntos, tramos = construir_tabla_alineamiento(vertices)
    perfil_puntos, perfil_tramos = construir_perfil_longitudinal(
        puntos,
        tramos,
        [100.0, 104.0, 101.0],
    )

    _, tabla_tramos = serializar_tablas_perfil(perfil_puntos, perfil_tramos)

    assert tabla_tramos.filas[0][3] == pytest.approx(50.0)
    assert tabla_tramos.filas[0][4] == pytest.approx(4.0)
    assert tabla_tramos.filas[0][6] == pytest.approx(8.0)
    assert tabla_tramos.filas[1][4] == pytest.approx(-3.0)
